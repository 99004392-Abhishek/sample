from openpyxl import workbook, load_workbook

wb = load_workbook('Genesis mini_project.xlsx')

ws = wb.active

def my_ps_number(ip_ws):
    ps_nums=[]
    for row in ws.iter_rows(min_row=1, max_col=1, max_row=16, values_only=True):
        ps_nums.append(list(row))
    return ps_nums

def show_ps_number(ps):
    for item in ps:
        print(item)

def main():
    print("\nEnter the PS number from the below list:\n")
    ps = my_ps_number(ws)
    show_ps_number(ps)
    user_choice = int(input("\nEnter the Ps number:"))

main()